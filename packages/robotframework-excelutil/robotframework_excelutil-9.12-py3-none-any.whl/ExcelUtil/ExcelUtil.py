#!/usr/bin/env python


import openpyxl
import os

class ExcelUtil:
    """
        This test library internally use openpyxl module of python and provides keywords to open, read,
        write excel files. This library only supports xlsx file formats.


        *Prerequisties*

        1. Openpyxl module of python should be installed using command "pip install openpyxl"
        
        2. ExcelUtil written by Nagesh Nagaraja should be installed uding command 
        "pip install robotframework-excelutil"
        
        3. ExcelUtil must be imported.

        Example:
            | Library        | ExcelUtil        |
            | Open Excel     | Filename with fullpath |

        """

    def __init__(self):
        self.wb = None
        self.sheet = None
        self.filename = None
        self.fileDir = os.path.dirname(os.path.realpath('__file__'))

    def open_excel(self, file):
        """
        Open excel file
        Arguments:
            | File             | Filename with fullpath to open and test upon        |

        Example:
        | Open Excel      |  C:\\Data\\ExcelTest.xlsx  |
        """
        self.filename = file
        self.wb = openpyxl.load_workbook(self.filename)

    def open_excel_from_relative_path(self, file):
        """
        Open excel from relative path
        Arguments:
            | File             | Filename with relative path from project home directory 
                                 to open and test upon        |

        Example:
        | Open Excel From Relative Path     |  ../../Data/data.xlsx  |
        """
        self.filename = os.path.join(self.fileDir, file)
        self.wb = openpyxl.load_workbook(self.filename)

    def open_excel_from_project_directory(self, file):
        """
        Open excel from project directory
        Arguments:
            | File             | Filename from project folder to open and test upon        |

        Example:
        | Open Excel From Project Directory     |  data.xlsx  |
        """
        self.filename = os.path.join(self.fileDir, file)
        self.wb = openpyxl.load_workbook(self.filename)

    def get_sheet_names(self):
        """
        Return sheetnames of the workbook
        Example:
        | Open Excel      |  C:\\Data\\ExcelTest.xlsx  |
        | Get sheet names      |
        """
        return self.wb.get_sheet_names()

    def get_column_count(self, sheetname):
        """
        Return the column count of the given sheet
        Example:
        | Open Excel      |  C:\\Data\\ExcelTest.xlsx  |
        | Get Column count     |  Sheet1 |
        """
        # self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        return self.sheet.max_column

    def get_row_count(self, sheetname):
        """
        Return the Row count of the given sheet
        Example:
        | Open Excel      |  C:\\Data\\ExcelTest.xlsx  |
        | Get Row count     |  Sheet1 |
        """
        self.sheet = self.wb[sheetname]
        return self.sheet.max_row

    def get_sheet_count(self):
        """
        Return the Sheet count of the currently open Excel
        Example:
        | Open Excel      |  C:\\Data\\ExcelTest.xlsx  |
        | Get Sheet count     |
        """
        return len(self.wb.get_sheet_names())

    def read_cell_data_by_coordinates(self, sheetname, row_value, column_value):
        """
        Return the value of a cell by giving the sheetname, row value & column value
        Example:
        | Read Cell Data By Coordinates     |  SheetName | Row Number |  Column Number  |
        | Read Cell Data By Coordinates     |  Sheet1 |  1  |  1  |
         To pass integer arguments
        | Read Cell Data By Coordinates     |  Sheet1 |  ${1}  |  ${3}  |
        """
        # self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        varcellValue = self.sheet.cell(row=self.row, column=self.column).value
        return varcellValue


    def get_row_data_upto_column(self, sheetname, row_value, column_value):
        """
        Return the values of all cells in a row from column 1 to given column_value for given sheetname
        Example:
        | Get Row Data Upto Column     |  SheetName | Row Number |  Column Number  |
        | Get Row Data Upto Column     |  Sheet1 |  1  |  9  |
        """
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)

        self.rowdata = []

        if int(column_value) < 1:
            raise Exception('Column value cannot be less than 1')

        for i in range(1, int(column_value)+1):
            self.column = int(i)
            varcellValue = self.sheet.cell(row=self.row, column=self.column).value
            self.rowdata.append(varcellValue)
        return self.rowdata

    def get_all_rows_data(self, sheetname):
        """
        Return the values of all cells in all rows for given sheetname
        Example:
        | Get All Rows Data     |  SheetName |
        | Get All Rows Data     |  Sheet1 |
        """
        self.sheet = self.wb[sheetname]

        rows_data = []
        for row in self.sheet.iter_rows():
            rows_data.append(list( cell.value or "" for cell in row ) )
        return rows_data

    def write_data_by_coordinates(self, sheetname, row_value, column_value, varvalue):
        """
        Write the value to a call using its co-ordinates
        Example:
        | Write Data By Coordinates    |  SheetName  | Row Number | Column Number |  Data  |
        | Write Data By Coordinates    | Sheet1 | 1 | 1 |  TestData  |
        """
        # self.sheet = self.wb.get_sheet_by_name(sheetname)
        self.sheet = self.wb[sheetname]
        self.row = int(row_value)
        self.column = int(column_value)
        self.varValue = varvalue
        self.sheet.cell(row=self.row, column=self.column).value = self.varValue

    def save_excel(self, file):
        """
        Save the excel file after writing the data.
        Example:
        Update existing file:

        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |

        Save in new file:
        | Openexcel File       |  C:\\Python27\\ExcelRobotTest\\ExcelRobotTest.xlsx  |
        | Save Excelfile       |  D:\\Test\\ExcelRobotNewFile.xlsx                   |
        """
        self.file = file
        self.wb.save(self.file)

    def add_new_sheet(self, varnewsheetname):
        """
        Add new sheet
        Arguments:
        | New sheetname        | The name of the new sheet to be added in the workbook     |

        Example:
        | Keywords             | Parameters                                       |
        | Add new sheet        | SheetName                                       |
        """
        self.newsheet = varnewsheetname
        self.wb.create_sheet(self.newsheet)