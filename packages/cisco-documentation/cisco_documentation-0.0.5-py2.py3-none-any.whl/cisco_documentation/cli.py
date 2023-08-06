import sys
import netmiko.ssh_exception
import yamlarg
import os
import shutil
import socket
import csv
from napalm import get_network_driver
from ntc_templates import parse
import json
import keyring
from napalm.base.helpers import canonical_interface_name
from joblib import Parallel, delayed
from datetime import datetime

#Set rich to be the default method for printing tracebacks.
from rich.traceback import install
install(show_locals=True)

#import asyncio
#from aiomultiprocess import Pool

def is_open(ip,port):
   s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
   try:
      s.connect((ip, int(port)))
      s.shutdown(2)
      return True
   except:
      return False


def setpass(service, username):
    import keyring
    import getpass
    keyring.set_password(service,
                         username,
                         getpass.getpass('Enter the ' + username + ' for ' + service + ': '))


def get_or_set_password(service, username):
    import keyring
    creds = keyring.get_password(service, username)
    if creds is None:
        setpass(service, username)
        creds = keyring.get_password(service, username)
    return creds


def split_interface(interface):
    try:
        num_index = interface.index(next(x for x in interface if x.isdigit()))
        str_part = interface[:num_index]
        num_part = interface[num_index:]
    except StopIteration:
        return ['', '']
    return [str_part, num_part]


def get_oui_dict(pkgdir):
    f_in = open(os.path.join(pkgdir, 'templates/wireshark_oui.txt'), 'r')
    oui = filter(None, (line.partition('#')[0].rstrip() for line in f_in))
    oui_dict = dict()
    for line in oui:
        part = line.partition('\t')
        if 'IEEE Registration Authority' not in part[2]:
            mac_prefix = part[0].replace(':', '').replace('.', '')
            if len(mac_prefix) == 6:
                oui_dict[mac_prefix] = part[2].replace('\t', ', ')
            else:
                if mac_prefix[0:6] not in oui_dict.keys():
                    oui_dict[mac_prefix[0:6]] = dict()
                oui_dict[mac_prefix[0:6]][part[0]] = part[2].replace('\t', ', ')
    return oui_dict


def mac_to_bits(mac_address):
    return int(mac_address.replace(':', '').replace('.', ''), 16)


def mac_subnet(mac_address, subnet):
    mac = mac_to_bits(mac_address)
    low = mac_to_bits(subnet.partition('/')[0])
    high = mac_to_bits(subnet.partition('/')[0]) + int('1' * (48 - int(subnet.partition('/')[2])), 2)
    if mac >= low and mac <= high:
        return True
    else:
        return False


def oui_lookup(mac_address, oui_dict):
    mac_address = mac_address.replace(':', '').replace('.', '')
    if mac_address[0:6] in oui_dict.keys():
        if type(oui_dict[mac_address[0:6]]) == str:
            return oui_dict[mac_address[0:6]]
        else:
            for subnet, company in oui_dict[mac_address[0:6]].items():
                if mac_subnet(mac_address, subnet):
                    return company
    else:
        return '(Unknown)'


def collect_sw_info(switch):
    sw_ip = switch['switch']
    device_info = dict()
    driver = get_network_driver(switch['driver'])

    while True:
        try:
            un = get_or_set_password(switch['switch'], 'username')
            pw = get_or_set_password(switch['switch'], 'password')
            device = driver(switch['switch'], un, pw,
                            optional_args={'global_delay_factor': 2, 'transport': switch['transport']})
            device.open()
            break
        except netmiko.ssh_exception.AuthenticationException:
            print('Authentication Failed.')
            setpass(switch['switch'], 'username')
            setpass(switch['switch'], 'password')

    device_info['facts'] = device.get_facts()
    device_info['full-config'] = device.get_config(full=True)
    device_info['config'] = device.get_config()
    device_info['vlans'] = device.get_vlans()
    device_info['arp'] = device.get_arp_table()
    device_info['mac'] = device.get_mac_address_table()
    device_info['users'] = device.get_users()
    device_info['interfaces'] = device.get_interfaces()
    device_info['lldp'] = device.get_lldp_neighbors_detail()
    if switch['driver'] == 'ios' and switch['transport'] == 'ssh':
        device_info['cdp'] = device.cli(['show cdp neighbors detail'])['show cdp neighbors detail']
        device_info['cdp-parsed'] = parse.parse_output('cisco_ios', 'show cdp neighbors detail', device_info['cdp'])
        device_info['trans'] = device.cli(['show int trans'])['show int trans']
        device_info['int'] = device.cli(['show interfaces'])['show interfaces']
        device_info['int-parsed'] = parse.parse_output('cisco_ios', 'show interfaces', device_info['int'])
    return [sw_ip, device_info]


def main():
    try:
        pkgdir = sys.modules['cisco_documentation'].__path__[0]
    except KeyError:
        import pathlib
        pkgdir = pathlib.Path(__file__).parent.absolute()

    oui_dict = get_oui_dict(pkgdir)

    args = yamlarg.parse(os.path.join(pkgdir, 'arguments.yaml'))

    if args['update_wireshark_oui']:
        import requests
        url = 'https://gitlab.com/wireshark/wireshark/raw/master/manuf'
        myfile = requests.get(url)
        open(os.path.join(pkgdir, 'templates/wireshark_oui.txt'), 'wb').write(myfile.content)

    if args['load_creds'] != '':
        with open(args['load_creds'], 'r') as csvfile:
            creds = csv.DictReader(csvfile, fieldnames=['ip', 'username', 'password'], delimiter=',')
            for cred in creds:
                keyring.set_password(cred['ip'], 'username', cred['username'])
                keyring.set_password(cred['ip'], 'password', cred['password'])
        os.remove(args['load_creds'])

    if args['excel_template']:
        dstfile = './Customer City, ST - IP Address Listing.xlsx'
        pkgfile = 'templates/Customer City, ST - IP Address Listing.xlsx'
        if not os.path.isfile(dstfile):
            shutil.copy(os.path.join(pkgdir, pkgfile), dstfile)

    if args['fetch_info']:
        info = dict()
        with open(args['switch_list'], 'r') as csvfile:
            switches = csv.DictReader(csvfile, fieldnames=['switch', 'driver', 'transport'], delimiter=',')
            next(switches)
            switch_list = [switch for switch in switches]
            # async with Pool() as pool:
            #     results = await pool.map(collect_sw_info, switch_list)
            results = Parallel(n_jobs=len(switch_list), verbose=0, backend='threading')(
                map(delayed(collect_sw_info), switch_list))
            for result in results:
                ip = result[0]
                device_info = result[1]
                info[ip] = device_info
        with open(os.path.join(args['output_dir'], 'output.json'), 'w') as f:
            f.write(json.dumps(info))

    if args['parse_info']:
        with open(os.path.join(args['output_dir'], 'output.json'), 'r') as f:
            info = json.loads(f.read())
        output_dict = dict()
        output_arp = list()
        for sw_ip, device_info in info.items():
            output_dict[sw_ip] = dict()
            # Save config files.
            sw_hostname = device_info['facts']['hostname']
            filename = os.path.join(args['output_dir'], sw_ip + '-' + sw_hostname)
            with open(filename + '-running.txt', 'w') as f:
                f.write(device_info['config']['running'])
            with open(filename + '-startup.txt', 'w') as f:
                f.write(device_info['config']['startup'])
            output_dict[sw_ip]['hostname'] = sw_hostname
            output_dict[sw_ip]['interfaces'] = dict()
            for interface,interface_values in device_info['interfaces'].items():
                output_dict[sw_ip]['interfaces'][interface] = dict()
                if interface in device_info['lldp']:
                    output_dict[sw_ip]['interfaces'][interface]['devices'] = ''
                    output_dict[sw_ip]['interfaces'][interface]['mac'] = ['']
                    output_dict[sw_ip]['interfaces'][interface]['neighbor'] = ''
                    for neighbor in device_info['lldp'][interface]:
                        output_dict[sw_ip]['interfaces'][interface]['neighbor'] += neighbor['remote_system_name'] + ' - ' + neighbor['remote_port']
                else:
                    output_dict[sw_ip]['interfaces'][interface]['devices'] =len([True for i in device_info['mac'] if canonical_interface_name(i['interface']) == interface])
                    output_dict[sw_ip]['interfaces'][interface]['mac'] = [i['mac'] for i in device_info['mac'] if canonical_interface_name(i['interface']) == interface]
                    if output_dict[sw_ip]['interfaces'][interface]['mac'] == []:
                        output_dict[sw_ip]['interfaces'][interface]['mac'] = ['']
                    output_dict[sw_ip]['interfaces'][interface]['neighbor'] = ''
                output_dict[sw_ip]['interfaces'][interface]['description'] = device_info['interfaces'][interface]['description']
                output_dict[sw_ip]['interfaces'][interface]['enabled/up'] = str(device_info['interfaces'][interface]['is_enabled']) + '/' + str(device_info['interfaces'][interface]['is_up'])
                output_dict[sw_ip]['interfaces'][interface]['speed'] = device_info['interfaces'][interface]['speed']
                output_dict[sw_ip]['interfaces'][interface]['duplex'] = [i['duplex'] for i in device_info['int-parsed'] if i['interface'] == interface]
                output_dict[sw_ip]['interfaces'][interface]['vlans'] = list()
                for vlan in device_info['vlans']:
                    if interface in device_info['vlans'][vlan]['interfaces']:
                        output_dict[sw_ip]['interfaces'][interface]['vlans'].append(vlan)
                for entry in device_info['arp']:
                    if 'mac' in entry.keys() and 'ip' in entry.keys():
                        output_arp.append([entry['ip'],
                                           entry['mac'].upper().replace(':','').replace('.',''),
                                           oui_lookup(entry['mac'], oui_dict)])
        with open(os.path.join(args['output_dir'], 'output.csv'), 'w') as f:
            f.write('name\tip\tint\tdevices\tdescription\tenabled/up\tneighbor\tspeed\tduplex\tmac\tvlans\n')
            for sw_ip, device_info in output_dict.items():
                for interface, int_info in device_info['interfaces'].items():
                    device = 0
                    for mac in int_info['mac']:
                        # name,ip,int,devices,description,enabled/up,neighbor,speed,duplex,mac,vlans
                        output = [device_info['hostname'],
                                  sw_ip,
                                  interface,
                                  str(device),
                                  int_info['description'],
                                  int_info['enabled/up'],
                                  int_info['neighbor'],
                                  str(int_info['speed']),
                                  '' if int_info['duplex'] == [] else int_info['duplex'][0],
                                  mac.replace(':', ''),
                                  ','.join(int_info['vlans'])]
                        f.write('\t'.join(output) + '\n')
                        device += 1
        with open(os.path.join(args['output_dir'], 'arp_output.csv'), 'w') as f:
            f.writelines(','.join(entry) for entry in output_arp)

        if args['update_excel'] != '':
            from openpyxl import load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            wb = load_workbook(args['update_excel'])
            del wb['SWITCHES']
            ws = wb.create_sheet('SWITCHES')
            ws.append(['', 'SWITCH', 'SW IP', 'INT', 'DEVICE', 'DESCRIPTION',
                       'LINE PROTO', 'NEIGHBOR & PORT', 'SPEED', 'DUPLEX', 'MAC',
                       'VLAN', 'IP LOOKUP', 'NETWORK', 'INTEGRATOR',
                       'DEVICE / APPLICATION', 'DEVICE DESCRIPTION',
                       'DEVICE NAME'])

            for sw_ip, device_info in output_dict.items():
                for interface, int_info in device_info['interfaces'].items():
                    device = 0
                    for mac in int_info['mac']:
                        # name,ip,int,devices,description,enabled/up,neighbor,speed,duplex,mac,vlans
                        output = ['', # first column is left for navigation links.
                                  device_info['hostname'],
                                  sw_ip,
                                  interface,
                                  str(device),
                                  int_info['description'],
                                  int_info['enabled/up'],
                                  int_info['neighbor'],
                                  str(int_info['speed']),
                                  '' if int_info['duplex'] == [] else int_info['duplex'][0],
                                  mac.replace(':', ''),
                                  ','.join(int_info['vlans'])]
                        ws.append(output)
                        device += 1
            for row in range(2, ws.max_row):
                ws.cell(row=row, column=13,
                        value='=IFERROR(INDEX(ARP!A:A,MATCH(SWITCHES!K' + str(row) + ',ARP!B:B,0)),"")')
                ws.cell(row=row, column=13,
                        value='=IFERROR(INDEX(OVERVIEW!B:B,MATCH(L' + str(row) + ',OVERVIEW!D:D,0)),"")')
                ws.cell(row=row, column=13,
                        value='=IF(M' + str(row) + '<>"",INDEX(INDIRECT(N' + str(row) + '&"!B1:B99999"),MATCH(M' + str(row) + ',INDIRECT(N' + str(row) + '&"!F1:F99999"),0)),"")')
                ws.cell(row=row, column=13,
                        value='=IF(M' + str(row) + '<>"",INDEX(INDIRECT(N' + str(row) + '&"!C1:C99999"),MATCH(M' + str(row) + ',INDIRECT(N' + str(row) + '&"!F1:F99999"),0)),"")')
                ws.cell(row=row, column=13,
                        value='=IF(M' + str(row) + '<>"",INDEX(INDIRECT(N' + str(row) + '&"!D1:D99999"),MATCH(M' + str(row) + ',INDIRECT(N' + str(row) + '&"!F1:F99999"),0)),"")')
                ws.cell(row=row, column=13,
                        value='=IF(M' + str(row) + '<>"",INDEX(INDIRECT(N' + str(row) + '&"!E1:E99999"),MATCH(M' + str(row) + ',INDIRECT(N' + str(row) + '&"!F1:F99999"),0)),"")')

            ws['A1'].hyperlink = 'OVERVIEW!A1'
            ws['A1'].value = 'OVERVIEW'
            ws['A1'].style = 'Hyperlink'
            ws['A2'].hyperlink = "'OVERVIEW SWITCHES'!A1"
            ws['A2'].value = 'OVERVIEW SWITCHES'
            ws['A2'].style = 'Hyperlink'
            tab = Table(displayName="SWITCHES", ref="B1:R" + str(ws.max_row))
            style = TableStyleInfo(name="Table Style Light 1",
                                   showFirstColumn=False,
                                   showLastColumn=False,
                                   showRowStripes=True,
                                   showColumnStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)
            col_widths = {'A': 120,
                          'B': 160,
                          'C': 90,
                          'D': 123,
                          'E': 54,
                          'F': 260,
                          'G': 84,
                          'H': 360,
                          'I': 50,
                          'J': 65,
                          'K': 85,
                          'L': 53,
                          'M': 90,
                          'N': 90,
                          'O': 120,
                          'P': 150,
                          'Q': 250,
                          'R': 120}
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width / 6

            ws = wb['ARP']
            ws.append([datetime.now()])
            for entry in output_arp:
                ws.append(entry)
            wb.save(args['update_excel'])


