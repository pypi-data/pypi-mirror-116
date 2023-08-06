""" This module create a model to create a schema, parse data and import data into Weaviate """

import json
import fnmatch
import openpyxl

from .schema import Schema
from .data import Data
from .classification import Classification

from .utilities import get_weaviate_client
from .utilities import get_class_name_from_key

from .exceptions import UnableToGetWeaviateClient
from .exceptions import UnableToOpenModelFile
from .exceptions import UnsupportedModelType
from .exceptions import UnknownModelType


def _read_model_json(node: dict, current: dict, model: dict) -> dict:
    #pylint: disable=too-many-branches
    """
    Read the data model file from an json file. Note this is a recursive function

    Parameters
    ----------
    node : dict
        the json node that we are processing in this recursive iteration
    current : dict
        the current class in the model
    model : dict
        The dict containing the model

    Returns
    ------
    dict
        The dict describing the data model
    """

    if model is None:
        model = {}

    for key in node:
        if isinstance(node[key], (list, dict)):
            classname = get_class_name_from_key(key)
            if current is not None:
                current['of'+classname] = classname

            if classname not in model:
                model[classname] = {}

            for item in node[key]:
                if isinstance(node[key], list):
                    _read_model_json(item, model[classname], model)
                else:
                    _read_model_json(node[key], model[classname], model)

        elif isinstance(node[key], str):
            current[key] = 'string'

        elif isinstance(node[key], int):
            current[key] = 'int'

        elif isinstance(node[key], float):
            current[key] = 'number'

        elif isinstance(node[key], bool):
            current[key] = 'boolean'

        else:
            print(key, "something else")

    return model


def _read_model_excel(sheet: openpyxl.worksheet) -> dict:
    """
    Read the data model file from an excel file

    Parameters
    ----------
    sheet : openpyxl.worksheet
        The excel worksheet that contains the model

    Returns
    ------
    dict
        The dict describing the data model
    """

    # Initialize the return value
    model = {}
    model['classes'] = []

    # process the data line by line
    end = False
    row = 1
    while not end:

        if isinstance(sheet.cell(row=row, column=1).value, str):
            newclass = {}
            newclass['classname'] = sheet.cell(row=row, column=1).value
            newclass['columns'] = []
            model['classes'].append(newclass)

        elif isinstance(sheet.cell(row=row, column=1).value, int):
            number = int(sheet.cell(row=row, column=1).value)
            column = {}
            column['number'] = number
            name = sheet.cell(row=row, column=2).value
            if name.startswith("id:"):
                column['name'] = name[3:]
                column['id'] = True
            else:
                column['name'] = name
                column['id'] = False

            column['type'] = sheet.cell(row=row, column=3).value
            column['entity'] = bool(sheet.cell(row=row, column=4).value)
            column['indexInverted'] = bool(sheet.cell(row=row, column=5).value)

            newclass['columns'].append(column)

        elif sheet.cell(row=row, column=1).value is None:
            end = True

        row += 1

    return model


def _read_model(path):

    model = {}
    model['model'] = None
    model['type'] = ""

    if fnmatch.fnmatch(path, "*.xls") or fnmatch.fnmatch(path, "*.xlsx"):
        workbook = openpyxl.load_workbook(path, data_only=True)
        if workbook is not None:
            sheet = workbook.active
            if sheet is not None:
                model['model'] = _read_model_excel(sheet)
                model['type'] = "excel"

    elif fnmatch.fnmatch(path, "*.json"):
        with open(path) as jsonfile:
            root = json.load(jsonfile)
            if root is not None:
                model['model'] = _read_model_json(root, None, None)
                model['type'] = "json"

    elif fnmatch.fnmatch(path, "*.cvs"):
        raise UnsupportedModelType("csv")
    elif fnmatch.fnmatch(path, "*.yaml") or fnmatch.fnmatch(path, "*.yml"):
        raise UnsupportedModelType("yaml")
    else:
        raise UnknownModelType()

    if model['model'] is None or model['model'] == {}:
        raise UnableToOpenModelFile()

    return model


#########################################################################################################
# The class "Model"
#########################################################################################################


class Model:
    #pylint: disable=too-few-public-methods
    """
    Class that holds a data model to create schema, parse data and import data into Weaviate
    """

    def __init__(self, path: str, instance: dict, classification: dict=None):
        """
        Read the data model file.

        Parameters
        ----------
        path : str
            The path where the data model can be found

        Raises
        ------
        TypeError
            If arguments are of a wrong data type;
        UnknownModelType
            If the modeltype is unknown
        UnsupportedModelType
            If the modeltype is not supported yet
        """

        if not isinstance(path, str):
            raise TypeError("path is expected to be dict but is " + str(type(path)))
        if not isinstance(instance, dict):
            raise TypeError("instance is expected to be dict but is " + str(type(instance)))

        self.instance = instance
        self.client = get_weaviate_client(self.instance)

        model = _read_model(path)
        self.model = model['model']
        self.type = model['type']

        self.schema = Schema(model, self.instance, classification)
        self.data = Data(model, self.instance)
        self.classification = Classification(model, self.instance, classification)


    def get_client(self):
        """
        Returns the Weaviate client indicated by the argument instance

        Parameters
        ----------
        instance: dict
            A dict that contains all the weaviate parameters

        Raises
        ------
        UnableToGetWeaviateClient
            if we are unable to get the Weaviate client

        Returns
        -------
        weaviate:client
            the weaviate client indicated to by the argument dict instance
        """

        if self.client is None:
            raise UnableToGetWeaviateClient()

        return self.client
